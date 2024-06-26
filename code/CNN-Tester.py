# Benjamin E Hogan
# 01/29/2024

# --- Testing and Obtaining Results from Neural Network --- # 


# -------- Modules --------- #
import tensorflow as tf
import numpy as np 
import os
from tqdm import tqdm
import matplotlib.pyplot as plt
from tensorflow.keras.layers import Input
from tensorflow.keras.layers import Conv1D, MaxPooling1D, UpSampling1D
import pandas as pd
from tensorflow.keras.models import Model
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image as ReportLabImage
from reportlab.lib import colors
from reportlab.lib.units import inch

# -------- Global Vairables --------- #
BASE_DIR = os.getcwd()

# Directory paths
directories = {
    #Creates a dictionary to store paths of various directories
    "Pre-normalized": os.path.join(BASE_DIR, 'data', 'Pre-normalized_Files'),
    "normalized": os.path.join(BASE_DIR, 'data', 'Normalized_Files'),
    "train": os.path.join(BASE_DIR, 'data', 'data_set', 'train'),
    "val": os.path.join(BASE_DIR, 'data', 'data_set', 'val'),
    "test": os.path.join(BASE_DIR, 'data', 'data_set', 'test'),
    "models": os.path.join(BASE_DIR, "data", "Saved Models"),
    "data": os.path.join(BASE_DIR, "data", "Strain data"),
    "nan_files": os.path.join(BASE_DIR, "data", "NaN_Files"),
    "raw_data": os.path.join(BASE_DIR, "data", "Raw URL Data"),
    "spectrograms": os.path.join(BASE_DIR, "data", "images", "Spectrograms"),
    "Anomalies": os.path.join(BASE_DIR, "data", "images", "Anomalies"),
    "Reconstructions": os.path.join(BASE_DIR, "data", "images", "Reconstructions"),
    "images": os.path.join(BASE_DIR, "data", "images"),
    "Comparison": os.path.join(BASE_DIR, "data", "images", "Comparisons"),
    "Reports": os.path.join(BASE_DIR, "Reports")
}


# ------- Helper Functions -------- #

def load_data(directory, batch_size=24):
    print(f"Listing files in {directory}...")
    file_paths = [os.path.join(directory, f) for f in os.listdir(directory) if f.endswith('.npy')]

    print(f"Loading data from {directory}...")
    dataset = tf.data.Dataset.from_tensor_slices(file_paths)

    def load_file(file_path):
        data = np.load(file_path.numpy())
        data = np.expand_dims(data, axis=-1)
        return data, file_path  # Return both data and path

    dataset = dataset.map(lambda x: tf.py_function(load_file, [x], [tf.float32, tf.string]), num_parallel_calls=tf.data.experimental.AUTOTUNE)
    dataset = dataset.batch(batch_size).prefetch(tf.data.experimental.AUTOTUNE)
    print("All data loaded and batched")
    return dataset

    # Map the dataset through the load_file function, converting to TensorFlow objects
    dataset = dataset.map(lambda x: tf.py_function(load_file, [x], [tf.float32, tf.float32]), num_parallel_calls=tf.data.experimental.AUTOTUNE)

    print(f"Batching and prefetching...")
    # Batch and prefetch the dataset for optimal performance during training
    dataset = dataset.batch(batch_size).prefetch(tf.data.experimental.AUTOTUNE)
    print("All data loaded and batched")

    return dataset

def display_reconstructions(model, dataset, n=10):
    """
    Displays original and reconstructed data from a dataset for a specified number of samples.

    Parameters:
    - model: The trained model used for predictions.
    - dataset: The dataset from which data is taken for reconstruction.
    - n (int, optional): The number of data samples to display. Default is 10.
    """
    # Shuffle the dataset and take 1 batch
    for test_images, _ in dataset.shuffle(1000).take(1):
        decoded_images = model.predict(test_images)

        # Set up the figure for plotting
        fig, axes = plt.subplots(2, n, figsize=(20, 4))
        time = np.linspace(0, 64, 262144)  # Define a time axis for plotting

        # Determine global min and max for consistent plotting scales
        global_min = min(test_images.numpy().min(), decoded_images.min())
        global_max = max(test_images.numpy().max(), decoded_images.max())

        for i in range(n):
            if i >= test_images.shape[0]:  # Check to avoid out-of-bounds error
                break
            #Covert to LIGO
            # Plot original data
            ax = axes[0, i]
            ax.plot(time, test_images[i].numpy().flatten(), color='black')
            ax.set_title("Original")
            ax.set_xlabel('Time (s)')
            ax.set_ylabel('Strain')
            ax.set_ylim(global_min, global_max)

            # Plot reconstructed data
            ax = axes[1, i]
            ax.plot(time, decoded_images[i].flatten(), color='red')
            ax.set_title("Reconstructed")
            ax.set_xlabel('Time (s)')
            ax.set_ylabel('Strain')
            ax.set_ylim(global_min, global_max)

        plt.tight_layout()
        plt.show()

def create_directories(directory_paths):
    """
    Creates directories from a given dictionary of paths, if they do not already exist.

    Parameters:
    - directory_paths (dict): A dictionary where keys are directory names and values are paths.
    """
    for path in directory_paths.values():
        # Create each directory if it doesn't exist, avoiding errors if it already exists
        os.makedirs(path, exist_ok=True)

def save_signal_as_image(signal, file_path):
    """
    Saves a 1D signal as an image by plotting it and saving the plot.

    Parameters:
    - signal: The 1D signal array.
    - file_path: Path where to save the image.
    """
    plt.figure(figsize=(10, 2))  # Adjust size as needed
    plt.plot(signal)
    plt.axis('off')  # Hide axis
    plt.savefig(file_path, bbox_inches='tight', pad_inches=0)
    plt.close()  # Close the figure to free memory

def insert_images_to_excel(excel_path, img_col='Plot Path'):
    """
    Insert images into an Excel file based on the paths provided in a specific column.
    
    Args:
    excel_path (str): Path to the Excel file.
    img_col (str): The column header in the Excel file that contains the image paths.
    """
    # Load the workbook and grab the active worksheet
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Determine the column number for img_col
    img_col_idx = None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == img_col:
            img_col_idx = col[0].column
            break
    
    if img_col_idx is None:
        print(f"Column {img_col} not found.")
        return
    
    # Insert images
    for row in range(2, ws.max_row + 1):  # Skipping header row
        cell_value = ws.cell(row=row, column=img_col_idx).value
        if cell_value and os.path.exists(cell_value):
            img = Image(cell_value)
            # Assuming you want to insert images in the next column
            img.anchor = ws.cell(row=row, column=img_col_idx + 1).coordinate
            ws.add_image(img)
            # Optionally, clear the cell with the path
            ws.cell(row=row, column=img_col_idx).value = None
    
    # Save the workbook
    wb.save(excel_path)

def extract_gps_time(file_path):
    """
    Extracts the GPS time from the file name, assuming it follows the keyword 'around' and consists of digits.

    Parameters:
    - file_path (str): The path of the file.

    Returns:
    - str: The extracted GPS time as a string, or None if not found.
    """
    # Regex pattern to find digits that follow the word 'around'
    match = re.search(r"around\s+(\d+)", file_path)
    if match:
        gps_time = match.group(1)
        print(f"GPS time extracted: {gps_time} from {file_path}")  # Debugging output
        return gps_time
    else:
        print(f"No GPS time found in {file_path}")  # Debugging output
        return None

def save_raw_predictions(reconstructed, gps_time, save_dir=directories["Reconstructions"]):
    """
    Saves the raw reconstructed predictions to a specified directory, named based on the GPS time.

    Parameters:
    - reconstructed (np.array): The reconstructed signal data.
    - gps_time (str): The GPS time extracted from the file name.
    - save_dir (str): The directory where the reconstructed predictions will be saved.
    """
    # Ensure the save directory exists
    os.makedirs(save_dir, exist_ok=True)
    
    # Define the file path for saving the reconstructed signal
    save_path = os.path.join(save_dir, f"reconstruction-{gps_time}.npy")
    
    # Save the numpy array to a file
    np.save(save_path, reconstructed)

# -------- Neural Network Functions -------- #

def load_latest_checkpoint(model, checkpoint_dir):
    """
    Loads the latest checkpoint from a directory into a model.

    Parameters:
    - model: The model to load the weights into.
    - checkpoint_dir (str): The directory where the model checkpoints are stored.

    Returns:
    - int: The epoch number of the last saved model if found, otherwise 0.
    """
    # List all .h5 model files in the checkpoint directory
    checkpoint_files = [os.path.join(checkpoint_dir, f) for f in os.listdir(checkpoint_dir) if f.endswith('.h5')]
    
    if checkpoint_files:
        # Find the latest checkpoint based on creation time
        latest_checkpoint = max(checkpoint_files, key=os.path.getctime)
        print(f"Loading weights from {latest_checkpoint}")
        model.load_weights(latest_checkpoint)
        
        # Extract the epoch number from the checkpoint filename
        epoch_num_match = re.search(r"_(\d+).h5$", latest_checkpoint)
        last_epoch = int(epoch_num_match.group(1)) if epoch_num_match else 0
        recent_epoch = last_epoch
        print(f'Loaded lasted model at epoch {last_epoch}')
        return last_epoch
    else:
        print("No checkpoint found.")
        return 0
  
def create_autoencoder(input_shape):
    """
    Constructs an autoencoder neural network for 1D signal data. Autoencoders are a type of artificial neural network 
    used to learn efficient codings of unlabeled data. The aim of an autoencoder is to learn a representation 
    (encoding) for a set of data, typically for the purpose of dimensionality reduction. This implementation focuses 
    on 1D signal data, such as time series or audio signals, utilizing 1D convolutional layers.

    The architecture of this autoencoder consists of two main parts: the encoder and the decoder. The encoder part 
    compresses the input into a smaller, dense representation, and the decoder part attempts to reconstruct the 
    input from this compressed representation.

    Args:
        input_shape (tuple): The shape of the input data, expected to be in the format (steps, 1) for 1D arrays. The 'steps' 
                             represents the dimensionality or length of the input signal, and '1' signifies that each step is 
                             a scalar (as opposed to a vector of features).

    Returns:
        Model: A compiled TensorFlow/Keras model of the autoencoder with the Adam optimizer and mean squared error loss.

    The model uses ReLU activations for the encoder and decoder intermediate layers to introduce non-linearity, 
    allowing it to learn more complex patterns in the data. The final layer uses a sigmoid activation function 
    to ensure the output values are in the range [0, 1], which is particularly useful if the input data has been 
    normalized to this range.

    This autoencoder can be used for various applications including but not limited to anomaly detection in time series, 
    noise reduction, and feature extraction.
    """
    
    # Input layer specifying the shape of input data
    input_img = Input(shape=input_shape)  
    
    # Encoder
    # Conv1D layer with 16 filters, kernel size of 3, using 'relu' activation and 'same' padding
    x = Conv1D(16, 3, activation='relu', padding='same')(input_img)
    # MaxPooling1D layer for downsampling the input by a factor of 2, using 'same' padding
    x = MaxPooling1D(2, padding='same')(x)
    # Another Conv1D layer, now with 8 filters, to further process the data
    x = Conv1D(8, 3, activation='relu', padding='same')(x)
    # Final MaxPooling1D in the encoder part for further downsampling
    encoded = MaxPooling1D(2, padding='same')(x)
    
    # Decoder
    # Conv1D layer that starts the decoding process, with 8 filters
    x = Conv1D(8, 3, activation='relu', padding='same')(encoded)
    # UpSampling1D layer to increase the dimensionality of the data back to its original size
    x = UpSampling1D(2)(x)
    # Conv1D layer with 16 filters to further refine the decoded signal
    x = Conv1D(16, 3, activation='relu', padding='same')(x)
    # Another UpSampling1D layer to ensure the output size matches the input size
    x = UpSampling1D(2)(x)
    
    # Output layer to reconstruct the input signal. Uses a sigmoid activation to ensure the output
    # values are between 0 and 1, assuming the input data was normalized to this range.
    decoded = Conv1D(1, 3, activation='sigmoid', padding='same')(x)

    # Creation of the autoencoder model, specifying the input and output layers
    autoencoder = Model(input_img, decoded)
    # Compiling the model with Adam optimizer and mean squared error as the loss function.
    # Additionally, it tracks mean squared error and mean absolute error as metrics for evaluation.
    autoencoder.compile(optimizer='adam', 
                        loss='mean_squared_error', 
                        metrics=[tf.keras.metrics.MeanSquaredError(), 
                                 tf.keras.metrics.MeanAbsoluteError()])

    return autoencoder

def evaluate_model(model, test_dataset):
    """
    Evaluates the model on the test dataset.

    Parameters:
    - model: The model to be evaluated.
    - test_dataset: The test dataset.

    Returns:
    - tuple: The test loss, test mean squared error, and test mean absolute error.
    """
    print("Evaluating model on test data...")
    test_loss, test_mse, test_mae = model.evaluate(test_dataset)
    print(f"Test Loss: {test_loss}, Test MSE: {test_mse}, Test MAE: {test_mae}")
    return test_loss, test_mse, test_mae
    
def flag_anomalies(model, dataset, n=10, threshold=0.001):
    anomaly_flags = []
    mse_scores = []
    plot_paths = []
    gps_times = []

    for data, paths in dataset.take(n):  # data and paths are batches
        reconstructed = model.predict(data)
        for i in range(data.shape[0]):
            original = data[i].numpy().flatten()
            recon = reconstructed[i].flatten()
            mse = np.mean((original - recon) ** 2)
            
            file_path = paths[i].numpy().decode('utf-8')  # Decode the file path
            gps_time = extract_gps_time(file_path)
            if gps_time:
                gps_times.append(gps_time)
                plot_path = save_comparison_plot(original, recon, i)
                plot_paths.append(plot_path)
                anomaly_flag = "yes" if mse > threshold else "no"
                anomaly_flags.append(anomaly_flag)
                mse_scores.append(mse)
            else:
                print(f"Failed to extract GPS time from {file_path}")

    return pd.DataFrame({
        "GPS Time": gps_times,
        "MSE Score": mse_scores,
        "Is Anomaly": anomaly_flags,
        "Plot Path": plot_paths
    })

def save_comparison_plot(original, reconstructed, idx, save_dir=directories["Comparison"]):
    """Save plots comparing original and reconstructed signals, and their difference."""
    # Ensure the save directory exists
    os.makedirs(save_dir, exist_ok=True)
    
    # Plot original and reconstructed signals
    fig, axs = plt.subplots(3, 1, figsize=(10, 6), sharex=True)
    time = np.linspace(0, 1, len(original))
    axs[0].plot(time, original, label="Original")
    axs[0].set_title('Original Signal')
    axs[1].plot(time, reconstructed, label="Reconstructed", color='r')
    axs[1].set_title('Reconstructed Signal')
    
    # Plot the difference
    difference = np.abs(original - reconstructed)
    axs[2].plot(time, difference, label="Difference", color='g')
    axs[2].set_title('Difference Signal')
    
    # Save the figure
    fig.tight_layout()
    plot_path = os.path.join(save_dir, f"comparison_{idx}.png")
    plt.savefig(plot_path)
    plt.close(fig)
    return plot_path

def generate_excel_report(anomaly_image_paths, reconstruction_image_paths):
    """
    Generates an Excel report containing details of anomalies detected.

    Parameters:
    - anomaly_image_paths (list): Paths to images classified as anomalies.
    - reconstruction_image_paths (list): Paths to reconstructed images of the samples.

    Returns:
    - None
    """
    # Ensure the lists have the same length; adjust your logic as needed based on your intent
    assert len(anomaly_image_paths) == len(reconstruction_image_paths), "Lists must have the same length."

    data = {
        "Anomaly Image Path": anomaly_image_paths,
        "Reconstructed Image Path": reconstruction_image_paths,
    }
    df = pd.DataFrame(data)
    excel_path = os.path.join(BASE_DIR, "anomaly_report.xlsx")
    df.to_excel(excel_path, index=False)
    print(f"Excel report saved at {excel_path}")

def generate_pdf_report(data_frame, pdf_path):
    doc = SimpleDocTemplate(pdf_path, pagesize=letter)
    elements = []

    table_data = [['GPS Time', 'MSE Score', 'Is Anomaly', 'Comparison Plot']]

    for index, row in data_frame.iterrows():
        gps_time = row['GPS Time']
        mse_score = row['MSE Score']
        is_anomaly = row['Is Anomaly']
        plot_path = row['Plot Path']

        if os.path.exists(plot_path):
            img = ReportLabImage(plot_path)
            img.drawHeight = 1 * inch
            img.drawWidth = 2 * inch
        else:
            img = "Image not found"

        table_data.append([gps_time, f"{mse_score:.2e}", is_anomaly, img])

    table = Table(table_data, colWidths=[1*inch, 1*inch, 1*inch, 3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    doc.build(elements)
    print(f"PDF report generated at: {pdf_path}")

# -------- Define and run __MAIN__ ------ # 

def main():
    # --- Ensure Folders Exist --- #
    create_directories(directories)
    
    # ---- Load Data ----- #
    test_data = load_data(directories["test"])
    train_data = load_data(directories["train"])  # Note: train_data loaded but not used in this snippet

    # --- Load Model ---- #
    input_shape = (262144, 1)  # Assuming 1D signal with 262144 steps
    autoencoder = create_autoencoder(input_shape)
    last_epoch = load_latest_checkpoint(autoencoder, directories['models'])
    #ModelPlot(model=autoencoder, grid=False, connection=True, linewidth=0.1)
    # It seems like you're trying to load weights directly after creating your model. Assuming this is intentional:
    autoencoder.load_weights(f'data/Saved Models/best_model_{last_epoch}.h5')
    
    # --- Evaluate Model ---- #
    #evaluate_model(autoencoder, test_data)
    #display_reconstructions(autoencoder, test_data, n=10)
    
    # Flag anomalies and generate report
    anomaly_report = flag_anomalies(autoencoder, test_data, n=10, threshold=0.001)
    
    # Optionally, save the report to an Excel file
    #report_path = os.path.join(directories["Reports"],  "anomaly_report.xlsx")
    #insert_images_to_excel(report_path, img_col='Plot Path')
    #print(f"Anomaly report saved to {report_path}")
    
    # Generate PDF report
    pdf_report_path = os.path.join(directories["Reports"], "anomaly_report.pdf")
    generate_pdf_report(anomaly_report, pdf_report_path)

if __name__ == "__main__":
    main()
